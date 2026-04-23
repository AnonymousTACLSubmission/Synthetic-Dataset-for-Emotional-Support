from openpyxl import load_workbook
from better_profanity import profanity

# Load the profanity word list (default list)
profanity.load_censor_words()

# Load the Excel file
file_path = "-"  # Replace with your file path
workbook = load_workbook(filename=file_path)
sheet = workbook.active  # Use the active sheet, or specify a sheet with workbook['SheetName']

# Iterate through rows and columns
for row in sheet.iter_rows():
    for cell in row:
       print(f"Checking cell {cell.coordinate}")
       if cell.value:  # Check if the cell is not empty
            cell_text = str(cell.value)  # Convert the cell value to a string
            if profanity.contains_profanity(cell_text):  # Check for profanity
                print(f"Profanity detected in cell {cell.coordinate}: {cell_text}")
            else:
                print("false")

       else:
            continue
